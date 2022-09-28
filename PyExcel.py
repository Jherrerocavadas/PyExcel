from openpyxl import Workbook, load_workbook
import tkinter as tk
import tkinter.ttk as ttk
import os
# O código necessita de revisões aprofundadas
from config_user_provisorio import Users
users = Users()



class Planilha():
    def __init__(self):
        #self.xlsx = None
        self.first_time = True
        self.sheet_is_selected = None




    def Criar_Arquivo(self,Title):
        self.xlsx = Workbook()
        self.plan = self.xlsx.active
        self.plan.title = Title

    def Carregar_Arquivo(self, path, first_time, table):#=None):
        #para reutilizar a função e carregar a tabela também

        if (path.endswith('.xlsx') != False):# and path != None):
            if(first_time == True):
                self.path = path
                self.xlsx = load_workbook(path)

                if (self.xlsx.sheetnames != []):
                    table['values'] = self.xlsx.sheetnames

                self.first_time = False
                print('Primeira vez')

            else:
                print('Table: ', table)
                try:

                    if(table.get() != None and table.get() != ''):
                        self.plan = self.xlsx[table.get()]
                        print('Segunda vez - if')
                        print(self.plan)
                    else:
                        self.plan = self.xlsx.active
                        print('Segunda vez - else')
                        print(self.plan)
                except KeyError:
                    print("KEY ERROR")
                finally:
                    self.first_time = True
                    print('Segunda vez - end')

        #





    def Inserir_Dados(self, Nome, Telefone,
    Telefone2, Telefone3, Email, Endereco, CEP, CPF_CNPJ):

        valores = [
        #("Nome", "Telefone", "Telefone 2", "Endereço", "CEP", "Email", "CPF/CNPJ"),
        (None, Nome, Telefone, Telefone2, Telefone3, Email, Endereco, CEP, CPF_CNPJ)]

        for linha in valores:
            self.plan.append(linha)

        self.Salvar()



    def Editar_Dados(self):
        self.plan['A7'] = "AOBA"
        self.plan['B5'] = 'BÊ cinco'
        self.Salvar()


    def Excluir_Dados(self):

        #self.Salvar()
        pass


    def Pesquisar_Dados(self):

        pass

    def Salvar(self):
        try:
            self.xlsx.save(self.path)
            print("Arquivo salvo!")
        except PermissionError:
            print('O arquivo do excel deve estar fechado para que as alterações possam ser salvas nele.')
            #print('')


# Modelo de classe que herda direto de tk.Tk()
class UI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Atualizador - Excel - V1.1")
        self.geometry("250x150")


#Criação dos elementos de Toplevel
        self.open_table = None
        self.insert_data = None
        self.edit_data = None
        self.exclude_data = None
        self.search_data = None



        self.PyExcel_dados = Planilha()
        #self.PyExcel_dados.Criar_Arquivo('Teste_excel1')
        self.estilos()
        self.menus(self)
        self.Inicio()

    #Seleção do Usuário
    def Inicio(self):

        def verify_login():
            # if(User.get() == 'users.user2' or User.get() == 'users.user1'):
            #     if(Senha.get() == "admin"):
            #         self.Carregar_Arquivo_UI(User.get())
            #     else:
            #         print('Not admin mode')
            # else:
            #     print('ERRO')

            self.Carregar_Arquivo_UI('users.user2')#User.get())

        frame_user = tk.Frame(self)
        frame_senha = tk.Frame(self)
        frame_entra1 = tk.Frame(self)

        frame_user.pack()
        frame_senha.pack()
        frame_entra1.pack()

        User_txt = tk.Label(frame_user, text = "Usuário:" )
        User = ttk.Entry(frame_user)

        User_txt.pack()
        User.pack()


        Senha_txt = tk.Label(frame_senha, text = "Senha:" )
        Senha = ttk.Entry(frame_senha)

        Senha_txt.pack()
        Senha.pack()

        Entrar = ttk.Button(frame_entra1, text = 'Entrar', command = verify_login)
        self.bind('<Return>',lambda event : verify_login())
        Entrar.pack()


    def Criar_Arquivo_UI(self,Title):
        self.PyExcel_dados.Criar_Arquivo()


    def Carregar_Arquivo_UI(self, User):
        self.withdraw_func(self.open_table)

        dirs = os.listdir(User)
        files = [file for file in dirs]

        if(self.open_table == None):
            self.open_table = tk.Toplevel(self)
            self.open_table.geometry("300x200")
            self.menus(self.open_table)
            Arquivos_txt = tk.Label(self.open_table, text = "Arquivos disponíveis:")
            Arquivos_txt.pack()

            Arquivos_select = ttk.Combobox(self.open_table, values = files)
            Arquivos_select.pack()

#bloquar mudanças dessa tabela durante a etapa
            Tabela_txt = tk.Label(self.open_table, text = "Tabelas na planilha selecionada:")
            Tabela_txt.pack()

            Tabela_select = ttk.Combobox(self.open_table)
            Tabela_select.pack()

            Entrar = ttk.Button(self.open_table, text = 'Selecionar',
             command = lambda : self.PyExcel_dados.Carregar_Arquivo(
             f'{User}\{Arquivos_select.get()}', self.PyExcel_dados.first_time, Tabela_select))
            Entrar.pack()

            self.open_table.bind('<Return>', lambda event : self.PyExcel_dados.Carregar_Arquivo(
            f'{User}\{Arquivos_select.get()}', self.PyExcel_dados.first_time, Tabela_select))

#  print( [Arquivos_select.get() if Arquivos_select.get() != None
# else Arquivos_select.get() for Arquivos_select.get() in Arquivos_select['values']])


        else:
            self.open_table.deiconify()

        self.open_table.lift()
        self.open_table.protocol("WM_DELETE_WINDOW", self.destroy)


    def Inserir_Dados_UI(self):
        self.withdraw_func(self.insert_data)
        if (self.insert_data == None):
            self.insert_data = tk.Toplevel(self)
            self.insert_data.geometry("500x500")
            self.menus(self.insert_data)

            frame_nome = tk.LabelFrame(self.insert_data, text ="Nome")
            frame_contatos = tk.LabelFrame(self.insert_data, text = "Contatos")
            frame_endereco = tk.LabelFrame(self.insert_data, text = "Endereço")
            frame_cpf_cnpj = tk.LabelFrame(self.insert_data, text = "CPF/CNPJ")
            frame_confirma = tk.Frame(self.insert_data)

            frame_nome.pack()
            frame_contatos.pack()
            frame_endereco.pack()
            frame_cpf_cnpj.pack()
            frame_confirma.pack()


            Nome = ttk.Entry(frame_nome)
            Telefone = ttk.Entry(frame_contatos)
            Telefone2 = ttk.Entry(frame_contatos)
            Telefone3 = ttk.Entry(frame_contatos)
            Email = ttk.Entry(frame_contatos)
            Endereco = ttk.Entry(frame_endereco)
            CEP = ttk.Entry(frame_endereco)
            CPF_CNPJ = ttk.Entry(frame_cpf_cnpj)
            Confirma = tk.Button(frame_confirma, text="Enviar", command = lambda : self.PyExcel_dados.Inserir_Dados(
            Nome.get(), Telefone.get(), Telefone2.get(), Telefone3.get(), Email.get(), Endereco.get(), CEP.get(), CPF_CNPJ.get()))

            Nome_txt = tk.Label(frame_nome, text='Nome: ')
            Telefone_txt = tk.Label(frame_contatos, text='Telefone: ')
            Telefone2_txt = tk.Label(frame_contatos, text='Telefone2: ')
            Telefone3_txt = tk.Label(frame_contatos, text='Telefone2: ')
            Email_txt = tk.Label(frame_contatos, text='Email: ')
            Endereco_txt = tk.Label(frame_endereco, text='Endereço: ')
            CEP_txt = tk.Label(frame_endereco, text='CEP: ')
            CPF_CNPJ_txt = tk.Label(frame_cpf_cnpj, text='CPF/CNPJ: ')

            Nome_txt.pack(side='left')
            Nome.pack()

            Telefone_txt.pack(side='left')
            Telefone.pack()

            Telefone2_txt.pack(side='left')
            Telefone2.pack()

            Telefone3_txt.pack(side='left')
            Telefone3.pack()

            Email_txt.pack(side='left')
            Email.pack()

            Endereco_txt.pack(side='left')
            Endereco.pack()

            CEP_txt.pack(side='left')
            CEP.pack()

            CPF_CNPJ_txt.pack(side='left')
            CPF_CNPJ.pack()

            Confirma.pack()

        else:
            self.insert_data.deiconify()

        self.insert_data.lift()
        self.insert_data.protocol("WM_DELETE_WINDOW", self.destroy)


    def Editar_Dados_UI(self):
        pass


    def Excluir_Dados_UI(self):

        #self.Salvar()
        pass


    def Pesquisar_Dados_UI(self):

        pass


    def menus(self,window_menu):
        menu = tk.Menu(window_menu)
        arquivos_menu = tk.Menu(menu,tearoff=0)
        tabela_menu = tk.Menu(menu,tearoff=0)
        consulta_menu = tk.Menu(menu,tearoff=0)

        tabela_menu.add_command(label ='Abrir Tabela', command=lambda: print('2'))
        tabela_menu.add_command(label ='Criar Tabela', command=lambda: print('2'))
        arquivos_menu.add_command(label = 'Inserir Dados', command = lambda: self.Inserir_Dados_UI())
        arquivos_menu.add_command(label = 'Editar Dados', command = lambda: print('2'))
        arquivos_menu.add_command(label = 'Excluir Dados', command = lambda: print('2'))
        consulta_menu.add_command(label = 'Pesquisar Dados', command = lambda: print('2'))
        consulta_menu.add_command(label = 'Salvar Dados', command = lambda: print('2'))
        menu.add_cascade(label ='Arquivo', menu=arquivos_menu)
        menu.add_cascade(label = 'Tabela', menu =tabela_menu)
        menu.add_cascade(label = 'Consulta', menu=consulta_menu)
        window_menu.config(menu=menu)

    def estilos(self):
        estilo = ttk.Style()
        estilo.configure('TEntry', padx = 5)

    def withdraw_func(self, aba_executada):
        #Os módulos são:
         # self (root)
         # self.open_table
         # self.insert_data
         # self.edit_data
         # self.exclude_data
         # self.search_data

         #tem que inserir todo módulo aqui manualmente
        abas = [self,
          self.open_table,
          self.insert_data,
          self.edit_data,
          self.exclude_data,
          self.search_data]
        abas.remove(aba_executada)

        for aba in abas:
            try:
                aba.withdraw()
            except AttributeError:
                pass
                # print('AE: ', aba)

if __name__ == '__main__':

    app = UI()
    #app.Inicio()

    # PyExcel_dados = Planilha()
    # #PyExcel_dados.Criar_Arquivo('Teste_excel1')
    # PyExcel_dados.Carregar_Arquivo("A1Excel1.xlsx")
    # PyExcel_dados.Inserir_Dados('Nome', 'Telefone', None, 'Endereco', 'CEP', 'Email', 'CPF_CNPJ')
    #
    # PyExcel_dados.Carregar_Arquivo("PyExcel/A1Excel.xlsx")
    # PyExcel_dados.Editar_Dados()
    #PyExcel_dados.xlsx.save("PyExcel\Arquivo Excel.xlsx")

    app.mainloop()
